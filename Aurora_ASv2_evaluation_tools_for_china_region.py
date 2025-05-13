import boto3
import traceback
import math
import json
import os
from datetime import datetime, timedelta
import time
import pandas as pd
import openpyxl
from openpyxl.chart import PieChart, Reference
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Font
import logging
import sys
from operator import itemgetter
import matplotlib.pyplot as plt
import numpy as np

IMG_WIDTH = 600
IMG_HEIGHT = 400
# 设置环境变量
#region_list = ['us-east-1','ap-northeast-1','us-east-2','us-west-1','us-west-2','ap-east-1','ap-south-1','ap-northeast-1','ap-northeast-2','ap-southeast-2','ca-central-1','eu-central-1','eu-west-1','eu-west-2','eu-west-3','eu-north-1','me-south-1','sa-east-1']

region_list = ['cn-north-1','cn-northwest-1']

print("Please select a region by entering the corresponding number:")
for i, rg in enumerate(region_list, start=1):
    print(f"{i}. {rg}")
user_input = input(f"Enter your choice (1-{len(region_list)}): ")
if user_input.isdigit() and 1 <= int(user_input) <= len(region_list):
    rds_region = region_list[int(user_input) - 1]
    print(f"You selected: {rds_region}")
else:
    print("Invalid input. Please try again.")

#rds_region = 'us-east-1'
os.environ['AWS_DEFAULT_REGION'] = rds_region

# 创建一个新的 Excel 工作簿
myworkbook = openpyxl.Workbook()

# 获取默认工作表
myworksheet = myworkbook.active
#myworksheet.title = "Summary"

# 获取当前日期时间作为日志文件名
log_filename = f"rds_explorer_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

# 配置logging
logging.basicConfig(
    filename=log_filename,
    level=logging.INFO,
    format='%(asctime)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

def get_asv2_price(rds_region):
    if "China (Beijing)" in rds_region:
        price_per_acu_hour = 0.98
    else:
        if "China (Ningxia)" in rds_region:
            price_per_acu_hour = 0.87
    return price_per_acu_hour

def get_instance_price(instance_class,rds_region,instance_vcpu,engine):
    aurora_price_list = {
        "db.r7g.large":{
            "engine": ["aurora mysql","aurora postgresql"],
            "China (Beijing)":{
                "od_price": 2.688,
                "1yri_np": 1.29,
                "3yri_ap": round(20498/36/730,4),  
            },
            "China (Ningxia)":{
                "od_price": 1.708,
                "1yri_np": 1.093,
                "3yri_ap": round(17949/36/730,4),
            }
        },
        "db.r6g.large":{
            "engine": ["aurora mysql","aurora postgresql"],
            "China (Ningxia)":{
                "od_price": 2.218,
                "1yri_np": 1.4269,
                "3yri_ap": round(23013/36/730,4)
            },
            "China (Beijing)":{
                "od_price": 2.616,
                "1yri_np": 1.2401,
                "3yri_ap": round(19937/36/730,4)
            }
        },
        "db.r5.large":{
            "engine": ["aurora mysql","aurora postgresql"],
            "China (Ningxia)":{
                "od_price": 2.4,
                "1yri_np": 1.544,
                "3yri_ap": round(24901/36/730,4)
            },
            "China (Beijing)":{
                "od_price": 2.89,
                "1yri_np": 1.37,
                "3yri_ap": "NA"
            }
        }
    } 

    rds_price_list = {
        "db.m5.large":{
            "engine": ["PostgreSQL","mysql"],
            "China (Ningxia)":{
                "od_price": 1.06,
                "1yri_np": 0.3704,
                "3yri_ap": round(6397/36/730,4) 
            },
            "China (Beijing)":{
                "od_price": 1.67,
                "1yri_np": 0.5858,
                "3yri_ap": "NA"
            }
        },
        "db.m5d.large":{
            "engine": ["PostgreSQL","mysql"],
            "China (Ningxia)":{
                "od_price": 2.028,
                "1yri_np": 1.5613,
                "3yri_ap": round(28241/36/730,4) 
            },
            "China (Beijing)":{
                "od_price": 2.199,
                "1yri_np": 1.6934,
                "3yri_ap": round(30631/36/730,4)
            }
        },
        "db.r5.large":{
            "engine": ["PostgreSQL","mysql"],
            "China (Ningxia)":{
                "od_price": 1.66,
                "1yri_np": 0.5824,
                "3yri_ap": round(8308/12/730,4)
            },
            "China (Beijing)":{
                "od_price": 2.4 ,
                "1yri_np": 0.8408,
                "3yri_ap": round(13258/12/730,4)
            }
        },
        "db.r5d.large":{
            "engine": ["PostgreSQL","mysql"],
            "China (Ningxia)":{
                "od_price": 2.617,
                "1yri_np": 2.0152,
                "3yri_ap": round(36453/12/730,4)
            },
            "China (Beijing)":{
                "od_price": 2.948,
                "1yri_np": 2.2703,
                "3yri_ap": round(41066/12/730,4)
            }
        },
        "db.m6g.large":{
            "engine": ["PostgreSQL","mysql"],
            "China (Ningxia)":{
                "od_price":0.9180,
                "1yri_np":0.3214,
                "3yri_ap": round(5551/12/730,4)
            },
            "China (Beijing)":{
                "od_price": 1.464,
                "1yri_np": 0.5125,
                "3yri_ap": round(8466/12/730,4)
            }
        },
        "db.r6g.large":{
            "engine": ["PostgreSQL","mysql"],
            "China (Ningxia)":{
                "od_price": 1.482,
                "1yri_np": 0.5187,
                "3yri_ap": round(7400/36/730,4)
            },
            "China (Beijing)":{
                "od_price": 2.15,
                "1yri_np": 0.7525,
                "3yri_ap": round(11866/36/730,4)
            }
        },
        "db.r6gd.large":{
            "engine": ["PostgreSQL","mysql"],
            #"China (Ningxia)":{
            #    "od_price": "NA",
            #    "1yri_np": "NA",
            #    "3yri_ap": "NA"
            #},
            "China (Beijing)":{
                "od_price": 2.641,
                "1yri_np": 2.0337,
                "3yri_ap": round(36788/36/730,4)
            }
        },
        #"db.c6gd.large":{
        #    "engine": ["PostgreSQL","mysql"],
        #    "China (Ningxia)":{ 
        #        "od_price": ,
        #        "1yri_np": ,
        #        "3yri_ap": 
        #    },
        #    "China (Beijing)":{
        #        "od_price": "NA",
        #        "1yri_np": "NA",
        #        "3yri_ap": "NA"
        #    }
        #},
        "db.r7g.large":{
            "engine": ["PostgreSQL","mysql"],
            "China (Ningxia)":{
                "od_price": 1.485,
                "1yri_np": 0.52,
                "3yri_ap": round(7805/36/730,4)
            },
            "China (Beijing)":{
                "od_price": 2.337,
                "1yri_np": 0.818,
                "3yri_ap": round(13508/36/730,4)
            }
        },
        "db.m7g.large":{
            "engine": ["PostgreSQL","mysql"],
            "China (Ningxia)":{
                "od_price": 0.939,
                "1yri_np": 0.319,
                "3yri_ap": round(5676/36/730,4)
            },
            "China (Beijing)":{
                "od_price": 1.647,
                "1yri_np": 0.576,
                "3yri_ap": round(9513/36/730,4)
            }
        }
    }

    factor = instance_vcpu/2
    instance_type = "db."+ instance_class.split('.')[1] + ".large"
    
    #print (instance_class,instance_type,rds_region,instance_vcpu,engine)
    if "aurora" in engine:
        base_line_price_list = aurora_price_list[instance_type][rds_region]
        #print ("base_line_price_list",base_line_price_list)
        od_price_per_unit = base_line_price_list['od_price']*factor
        noup_1yr_price_per_unit = base_line_price_list['1yri_np']*factor
        allup_3yr_price_per_unit = base_line_price_list['3yri_ap']*factor
    else :
        if engine == "mysql" or engine == "PostgreSQL":
            base_line_price_list = rds_price_list[instance_type][rds_region]
            #print ("base_line_price_list",base_line_price_list)
            od_price_per_unit = base_line_price_list['od_price']*factor
            noup_1yr_price_per_unit = base_line_price_list['1yri_np']*factor
            if base_line_price_list['3yri_ap'] != "NA":
                allup_3yr_price_per_unit = base_line_price_list['3yri_ap']*factor
            else:
                allup_3yr_price_per_unit = "NA"
    #print (od_price_per_unit,noup_1yr_price_per_unit,allup_3yr_price_per_unit)
    return (od_price_per_unit,noup_1yr_price_per_unit,allup_3yr_price_per_unit)

def get_instance_vcpu(instance_class):
    vcpu = {
        "large": 2,
        "xlarge": 4,
        "2xlarge": 8,
        "4xlarge": 16,
        "8xlarge": 32,
        "12xlarge": 48,
        "16xlarge": 64,
        "24xlarge": 96
    }
    instance_size = instance_class.split('.')[2]
    #print ("instance_size",instance_size)
    if instance_size in vcpu:
        vcpu_count = vcpu[instance_size]
        #print(f"{instance_class} 的 vCPU 数量为: {vcpu_count}")
        return vcpu_count
    else:
        logging.info(f"未找到 {instance_class} 对应的 vCPU 数量")

def get_cpu_utilization(instance_id, region='us-east-1', stat='Average'):
    """
    获取给定 RDS 实例在最近一周内的 CPU 使用率统计信息。

    参数:
    instance_id (str): RDS 实例的 ID
    region (str, optional): AWS 区域, 默认为 'us-east-1'
    stat (str, optional): 要获取的 CPU 使用率统计信息, 可选值为 'Average', 'Minimum', 'Maximum', 默认为 'Average'

    返回:
    float: CPU 使用率百分比
    """
    cloudwatch = boto3.client('cloudwatch', region_name=rds_region)

    response = cloudwatch.get_metric_data(
        MetricDataQueries=[
            {
                'Id': 'cpu',
                'MetricStat': {
                    'Metric': {
                        'Namespace': 'AWS/RDS',
                        'MetricName': 'CPUUtilization',
                        'Dimensions': [
                            {
                                'Name': 'DBInstanceIdentifier',
                                'Value': instance_id
                            }
                        ]
                    },
                    'Period': 86400*31,
                    'Stat': stat
                },
                'ReturnData': True
            }
        ],
        StartTime=datetime.utcnow() - timedelta(days=30),
        EndTime=datetime.utcnow()
    )

    #print(f"{instance_id}:{response['MetricDataResults']}")
    if response['MetricDataResults']:
        cpu_utilization = math.ceil(response['MetricDataResults'][0]['Values'][0])
        #print(cpu_utilization)
        return cpu_utilization
    else:
        return None


def get_aurora_serverless_acu_price(engine):
    """
    Retrieves the price per ACU-Hour for Aurora Serverless.

    Returns:
        float: The price per ACU-Hour for Aurora Serverless.
    """
    ## Q1:把 RDS 的 PG 和 MySQL 都替换为 Aurora PG 和 MySQL 是为了完成 RDS --> Aurora 的成本核算

    # Create a Pricing client 初始化 API
    pricing = boto3.client('pricing',region_name='us-east-1')
    if engine == "PostgreSQL":
        engine=engine.replace("PostgreSQL","Aurora PostgreSQL")
    if engine == "mysql":
        engine=engine.replace("mysql","Aurora MySQL")

    logging.info(f"acu price engine search: {engine}")
    # Define the parameters for the API request
    response = pricing.get_products(
        ServiceCode='AmazonRDS',
        Filters=[
            {'Type': 'TERM_MATCH','Field': 'productFamily','Value': 'ServerlessV2'},
            {'Type': 'TERM_MATCH', 'Field': 'location', 'Value': aws_region_to_location(rds_region)},
            {'Type': 'TERM_MATCH', 'Field': 'databaseEngine', 'Value': engine}
        ]
    )

    #product_json = json.loads(response['PriceList'][0])
    #logging.info(json.dumps(product_json, indent=4))
    # Extract the price information
    price_count = 0
    for price_list_item in response['PriceList']:
        product_json = json.loads(price_list_item)
        for offer_term_code, offer_term_data in product_json["terms"]["OnDemand"].items():
            for price_dimension_key, price_dimension_data in offer_term_data["priceDimensions"].items():
                price_per_unit = float(price_dimension_data["pricePerUnit"]["USD"])
                price_count += 1
                if price_count == 2:
                    #print ("price_per_unit",price_per_unit)
                    return price_per_unit

def count_cpu_usage_distribution(cpu_usage_data):
    """
    统计 CPU 使用率分布情况。

    参数:
    cpu_usage_data (list): 一个包含 CPU 使用率数据的列表。

    返回:
    dict: 一个包含每个 CPU 使用率范围及其对应数量的字典。
    """
    # 定义 CPU 使用率范围
    usage_ranges = [
        ('0% - 10%', 0, 10),
        ('10% - 20%', 10, 20),
        ('20% - 30%', 20, 30),
        ('30% - 50%', 30, 50),
        ('50% and above', 50, 100)
    ]

    # 初始化计数器
    usage_counts = [0] * len(usage_ranges)

    # 统计每个 CPU 使用率范围的数量
    for usage in cpu_usage_data:
        for i, (_, min_range, max_range) in enumerate(usage_ranges):
            if min_range <= usage < max_range:
                usage_counts[i] += 1
                break

    # 返回每个 CPU 使用率范围的数量
    result = [['CPU Usage Range', 'Percentage']]
    result.extend([[range_name, count] for range_name, count in zip(
        [range_name for range_name, _, _ in usage_ranges], usage_counts
    )])
    return result

def create_top_cpu_chart(data, worksheet, workbook):
    # 将数据写入工作表
    id=2
    worksheet.cell(row=1, column=12, value='instance')
    worksheet.cell(row=1, column=13, value='avg cpu')
    worksheet.cell(row=1, column=14, value='max cpu')
    instance_cpu_data = []
    for row in data:
        cells = row.split(',')
        #selected_columns = [0,1, 2, 4,6]
        selected_columns = [1,6,8]
        selected_cell = [cells[i] for i in selected_columns]
        for col, value in enumerate(selected_cell, start=1):
            if col == 2 or col == 3:
                value = int(value)
            worksheet.cell(row=id, column=12 + col - 1, value=value)
        instance_cpu_data = instance_cpu_data + selected_cell
        id = id +1

    instances = []
    avg_cpu_usage = []
    max_cpu_usage = []

    for i in range(0, len(instance_cpu_data), 3):
        instances.append(instance_cpu_data[i])
        avg_cpu_usage.append(int(instance_cpu_data[i+1]))
        max_cpu_usage.append(int(instance_cpu_data[i+2]))

    #创建柱状图的JPG图片
    img = Image(create_jpg_bar_chart(instances,avg_cpu_usage,max_cpu_usage))
    img.width = IMG_WIDTH
    img.height = IMG_HEIGHT


    # 将柱状图添加到工作表
    worksheet.add_image(img, "L16")

    # 保存 Excel 文件
    workbook.save("rds_report.xlsx")

def create_jpg_bar_chart(data1,data2,data3):
    #data = [['instance1', 20, 20], ['instance2', 30, 20], ['instance3', 25, 20], ...]
    # 将数据分离为实例名 AVG CPU  MAX CPU
    instances = data1
    avg_cpu_usage = data2
    max_cpu_usage = data3

    # 创建柱状图
    fig, ax = plt.subplots(figsize=(10, 4))
    bar_width = 0.4
    x = np.arange(len(instances))
    ax.bar(x - bar_width/2, avg_cpu_usage, bar_width, label='Avg CPU')
    ax.bar(x + bar_width/2, max_cpu_usage, bar_width, label='Max CPU')

    # 在柱子上显示数值
    for i, v in enumerate(avg_cpu_usage):
        ax.text(x[i] - bar_width/2, v, str(v), ha='center', va='bottom', fontsize=8)
    for i, v in enumerate(max_cpu_usage):
        ax.text(x[i] + bar_width/2, v, str(v), ha='center', va='bottom', fontsize=8)

    # 设置图表标题和坐标轴标签
    ax.set_title("Instance CPU Usage Top 8", fontsize=16)
    ax.set_xlabel("Instance", fontsize=14)
    ax.set_ylabel("CPU Usage (%)", fontsize=14)

    # 旋转 x 轴标签,以便更好地显示实例名称
    ax.set_xticks(x)
    ax.set_xticklabels(instances, rotation=20, fontsize=10)

    # 添加图例
    ax.legend(loc='upper right')

    # 调整图表边距,以确保 x 轴标签完全显示
    plt.subplots_adjust(bottom=0.3)

    # 保存图片为 JPG 格式
    file_name = "instance_avg_cpu_jpg_name.jpg"
    plt.savefig(file_name, dpi=300, bbox_inches='tight')
    return file_name

def create_top_cost_saving_chart(data, worksheet, workbook):
    # 将数据写入工作表
    id=2
    worksheet.cell(row=1, column=23, value='instance')
    worksheet.cell(row=1, column=24, value='Privision 1YR RI NP')
    worksheet.cell(row=1, column=25, value='ASv2 Cost 2/monthly')
    worksheet.cell(row=1, column=26, value='Save Percent 2')

    instance_cost_data = []
    for row in data:
        cells = row.split(',')
        selected_columns = [1,14,20,22]
        selected_cell = [cells[i] for i in selected_columns]
        for col, value in enumerate(selected_cell, start=1):
            if col == 2 or col == 3 or col ==4:
                value = float(value)
            worksheet.cell(row=id, column=23 + col - 1, value=value)
        instance_cost_data = instance_cost_data + selected_cell
        id = id +1

    instances = []
    ri_1year = []
    asv2_cost_2 = []
    asv2_2_saving_percent = []

    for i in range(0, len(instance_cost_data), 4):
        instances.append(instance_cost_data[i])
        asv2_cost_2.append(float(instance_cost_data[i+1]))
        ri_1year.append(float(instance_cost_data[i+2]))
        asv2_2_saving_percent.append(float(instance_cost_data[i+3]))
        #print (instances,asv2_cost_2,ri_1year,asv2_2_saving_percent)

    #创建柱状图的JPG图片
    img = Image(create_cost_bar_jpg(instances,asv2_cost_2,ri_1year))
    img.width = IMG_WIDTH
    img.height = IMG_HEIGHT


    # 将柱状图添加到工作表
    worksheet.add_image(img, "W16")

    # 保存 Excel 文件
    workbook.save("rds_report.xlsx")

def create_cost_bar_jpg(data1,data2,data3):
    #data = [['instance1', 20, 20], ['instance2', 30, 20], ['instance3', 25, 20], ...]
    # 将数据分离为实例名 EstCost  RI1Year
    instances = data1
    asv2_cost_2 = data2
    ri_1year = data3

    # 创建柱状图
    fig, ax = plt.subplots(figsize=(10, 4))
    bar_width = 0.4
    x = np.arange(len(instances))
    ax.bar(x - bar_width/2, asv2_cost_2, bar_width, label='Est Cost')
    ax.bar(x + bar_width/2, ri_1year, bar_width, label='1 Year RI')

    # 在柱子上显示数值
    for i, v in enumerate(asv2_cost_2):
        ax.text(x[i] - bar_width/2, v, str(v), ha='center', va='bottom', fontsize=8)
    for i, v in enumerate(ri_1year):
        ax.text(x[i] + bar_width/2, v, str(v), ha='center', va='bottom', fontsize=8)

    # 设置图表标题和坐标轴标签
    ax.set_title("Instance Cost Saving Top 8", fontsize=16)
    ax.set_xlabel("Instance", fontsize=14)
    ax.set_ylabel("Cost", fontsize=14)

    # 旋转 x 轴标签,以便更好地显示实例名称
    ax.set_xticks(x)
    ax.set_xticklabels(instances, rotation=20, fontsize=10)

    # 添加图例
    ax.legend(loc='upper right')

    # 调整图表边距,以确保 x 轴标签完全显示
    plt.subplots_adjust(bottom=0.3)

    # 保存图片为 JPG 格式
    file_name = "instance_cost_saving_jpg_name.jpg"
    plt.savefig(file_name, dpi=300, bbox_inches='tight')
    return file_name

def create_cpu_usage_distribution_chart(data, worksheet, workbook):
    """
    创建 CPU 使用率分布饼图并将其添加到 Excel 工作表中。

    参数:
    data (list): 包含 CPU 使用率分布数据的列表,第一行为列名。
    worksheet (Worksheet): 要添加图表的 Excel 工作表。
    workbook (Workbook): 包含工作表的 Excel 工作簿。
    """
    # 将数据写入工作表
    for row in data:
        worksheet.append(row)

    # 创建一个饼图
    #chart = PieChart()
    #chart.title = "Instance Count by CPU Avg Utilization"
    #chart.height = 12
    #chart.width = 18

    # 设置图表数据范围
    #labels = Reference(worksheet, min_row=2, max_row=len(data), min_col=1)
    #data_ref = Reference(worksheet, min_row=1, max_row=len(data), min_col=2)
    #chart.add_data(data_ref, titles_from_data=True)
    #chart.set_categories(labels)

    # 添加数字百分比标签
    #s = chart.series[0]
    #s.dLbls = DataLabelList()
    #s.dLbls.showVal = True

    # 创建饼图
    labels = [row[0] for row in data[1:] if float(row[1]) > 0]
    values = [float(row[1]) for row in data[1:] if float(row[1]) > 0]

    fig, ax = plt.subplots(figsize=(8, 8))
    patches, texts, autotexts = ax.pie(values, labels=labels, autopct='%1.1f%%')
    # 设置标签字体大小和颜色
    for t in texts:
        t.set_size('smaller')
        t.set_color('black')

    # 设置百分比字体大小和颜色
    for t in autotexts:
        t.set_size('smaller')
        t.set_color('white')

    ax.axis('equal')  # 确保饼图是圆形的
    ax.set_title("Instance Count by CPU Avg Utilization")

    # 将图像保存为 JPG 格式
    plt.savefig("cpu_usage_pie.jpg", dpi=300)

    # 将 JPG 图像插入到 Excel 工作表的 A16 单元格
    img = Image("cpu_usage_pie.jpg")
    img.width = IMG_WIDTH
    img.height = IMG_HEIGHT
    worksheet.add_image(img, "A16")

    # 保存 Excel 文件
    workbook.save("rds_report.xlsx")

def update_progress(current_step, total_steps):
    """
    更新进度条
    """
    progress = current_step / total_steps * 100
    bar_length = 30  # 进度条长度
    block = int(round(bar_length * progress / 100))
    text = "\rProcessing: [{0}] {1}%".format("#" * block + "-" * (bar_length - block), round(progress, 2))
    sys.stdout.write(text)
    sys.stdout.flush()

def aws_region_to_location(region):
    region_to_location_map = {
        "cn-north-1": "China (Beijing)",
        "cn-northwest-1": "China (Ningxia)"
    }

    if region in region_to_location_map:
        return region_to_location_map[region]
    else:
        return "Unknown location"

def main():
    output_result=[]
    avg_cpu_list=[]
    #mrr_output_result=[]
    output_result_chart=[]

    # MRR 1 年 RI total cost 变量
    #mrr_sum_cost_1ynp=0

    # MRR total asv2 cost 变量
    #mrr_sum_cost_asv2=0

    # 获取 RDS 实例列表
    rds = boto3.client('rds')
    response = rds.describe_db_instances()
    instance_list = response['DBInstances']
    p=0

    for instance in instance_list:
        instance_price_list = []
        update_progress(p + 1, len(instance_list))
        p=p+1
        logging.info("-----------------------")
        instance_id = instance['DBInstanceIdentifier']
        instance_class = instance['DBInstanceClass']
        engine = instance['Engine'].replace("-", " ", 1)
        engine_version = instance['EngineVersion']
        status = instance['DBInstanceStatus']
        availability_zone = instance['AvailabilityZone']
        db_cluster_identifier = instance.get('DBClusterIdentifier', '')
        account_id = instance['DBInstanceArn'].split(':')[4]

        #print ("instance_class",instance_class) 
        # 如果是 t3，t4，r4 系列的机型，则跳过
        #if all(cls not in instance_class for cls in ["t3", "t4", "r4","micro","small","medium","c6gd"]):
        #    pass
        #else:
        #    continue
        allowed_prefixes = {"r5.","r5d.","r6g.","r7g.","m5.","m5d.","m6g.","m7g.","r6gd."}
        if not any(prefix in instance_class for prefix in allowed_prefixes):
            continue
        else:
            pass

        if engine == 'aurora mysql' or engine == 'aurora postgresql' or engine == 'mysql' or engine == 'postgres':
            pass
        else:
            continue

        if instance_class == 'db.serverless':
            continue

        if engine == "postgres":
            engine = engine.replace("postgres", "PostgreSQL")

# 计算数据库实例使用的 storage 使用的情况
        volume_bytes_used_gb = 0
        if engine.startswith('aurora-'):
            cloudwatch = boto3.client('cloudwatch')
            response = cloudwatch.get_metric_statistics(
                Namespace='AWS/RDS',
                MetricName='VolumeBytesUsed',
                StartTime=datetime.utcnow() - timedelta(hours=4),
                EndTime=datetime.utcnow(),
                Period=3600,
                Statistics=['Average'],
                Dimensions=[
                    {
                        'Name': 'DBClusterIdentifier',
                        'Value': db_cluster_identifier
                    }
                ]
            )
            if response['Datapoints']:
                volume_bytes_used = response['Datapoints'][0]['Average']
                volume_bytes_used_gb = volume_bytes_used / (1024 ** 3)

        logging.info(f"Instance ID: {instance_id}")
        logging.info(f"Instance Type: {instance_class}")
        logging.info(f"Engine: {engine}-{engine_version}")
        logging.info(f"Status: {status}")
        logging.info(f"Availability Zone: {availability_zone}")
        
        logging.info(f"VolumeBytesUsed: {volume_bytes_used_gb:.2f} GB")
        #获取 instance vcpu 数量
        instance_vcpu = get_instance_vcpu(instance_class)
        #获取单价并计算 od，1年ri 无预付，以及 3年ri全预付成本
        instance_price_list = get_instance_price(instance_class,aws_region_to_location(rds_region),instance_vcpu,engine)
        od_price_per_unit = instance_price_list[0]
        od_monthly_cost = round(od_price_per_unit*730,3)
        noup_1yr_price_per_unit = instance_price_list[1]
        ri_1yr_no = round(noup_1yr_price_per_unit*730,3)
        allup_3yr_price_per_unit = instance_price_list[2]
        if allup_3yr_price_per_unit == "NA":
            ri_3yr_all = "NA"
        else:
            ri_3yr_all = round(float(allup_3yr_price_per_unit)*730,3)

        #获取acu的单价
        price_per_acu_hour = get_asv2_price(aws_region_to_location(rds_region))
        #print ("price_per_acu_hour",price_per_acu_hour) 

        # 获取最近1周的CPU 使用率情况 
        avg_cpu_util = get_cpu_utilization(instance_id, rds_region, stat='Average')
        #print ("avg_cpu_util",avg_cpu_util)
        
        if avg_cpu_util is not None:
            logging.info(f"Average CPU Utilization (1 week): {avg_cpu_util:.2f}%")
        else:
            logging.info("No CPU utilization data available")

        min_cpu_util = get_cpu_utilization(instance_id, rds_region, stat='Minimum')
        if min_cpu_util is not None:
            logging.info(f"Minimum CPU Utilization (1 week): {min_cpu_util:.2f}%")
        else:
            logging.info("No CPU utilization data available")

        max_cpu_util = get_cpu_utilization(instance_id, rds_region, stat='Maximum')
        if max_cpu_util is not None:
            logging.info(f"Maximum CPU Utilization (1 week): {max_cpu_util:.2f}%")
        else:
            logging.info("No CPU utilization data available")

        
        ## Q：下面的公式要除以 100 是为了将利用率转换为正常可计算的数值
        before_min_acu = (avg_cpu_util+min_cpu_util)/100/2*int(instance_vcpu)*4
        if before_min_acu <= 0.5:
            min_acu = 0.5
        else:
            min_acu = math.ceil((avg_cpu_util+min_cpu_util)/100/2*int(instance_vcpu)*4) 
        #print ("min_acu",min_acu)
        ## Q: 同上
        before_avg_acu=avg_cpu_util/100*int(instance_vcpu)*4
        if before_avg_acu <=0.5:
            avg_acu = 0.5
        else:
            avg_acu = math.ceil(avg_cpu_util/100*int(instance_vcpu)*4)
        #print('avg acu',avg_acu,"before_avg_acu",before_avg_acu,"avg_cpu_util",avg_cpu_util)
        # asv2 基于 avg cpu 的成本 
        asv2_cost_1 = avg_acu*730*price_per_acu_hour
        
        # 获取CPU 使用率超过MinACU的数据(=CPU%>(AVG+MIN)/2)
        sum_exceed_mincpu_cost = 0
        minacu_cpu=(avg_cpu_util+min_cpu_util)/2
        #print ("minacu_cpu",minacu_cpu)
        cloudwatch = boto3.client('cloudwatch', region_name=rds_region)
        response = cloudwatch.get_metric_data(
            MetricDataQueries=[
                {
                    'Id': 'm3',
                    'Expression': f'IF(m1>0,m1)',
                    'Label': 'CPU Watch'
                },
                {
                    'Id': 'm1',
                    'MetricStat': {
                        'Metric': {
                            'Namespace': 'AWS/RDS',
                            'MetricName': 'CPUUtilization',
                            'Dimensions': [
                                {
                                    'Name': 'DBInstanceIdentifier',
                                    'Value': instance_id
                                }
                            ]
                        },
                        'Period': 60,
                        'Stat': 'Maximum',
                        'Unit': 'Percent'
                    },
                    'ReturnData': False
                }
            ],
            StartTime=datetime.utcnow() - timedelta(days=30),
            EndTime=datetime.utcnow()
        )
        if response['MetricDataResults']:
            high_cpu_data = response['MetricDataResults'][0]
            timestamps = high_cpu_data['Timestamps']
            values = high_cpu_data['Values']
            timestamps.reverse()
            first_time = datetime.fromisoformat(str(timestamps[0])).strftime('%Y-%m-%d')
            last_time = datetime.fromisoformat(str(timestamps[-1])).strftime('%Y-%m-%d')
            values.reverse()
            all_values = values
            exceed_minacu_value_cnt = 0
            logging.info(f"all values cnt : {len(all_values)}")

            #计算弹性费用，超过MinACU的情况
            for i in range(len(timestamps)):
                if values[i] > minacu_cpu:
                    ## Q: 同上，values[i] 除以 100 是为了转换成可计算的数值
                    exceed_mincpu_cost = (math.ceil(values[i]/100*int(instance_vcpu)*4)-min_acu)*price_per_acu_hour/60
                    sum_exceed_mincpu_cost = sum_exceed_mincpu_cost + exceed_mincpu_cost
                    exceed_minacu_value_cnt = exceed_minacu_value_cnt + 1

            logging.info(f"CPU Utilization > {minacu_cpu}%: total {exceed_minacu_value_cnt}, only logging.info 10 samples")
            p_cnt = 0

            for i in range(len(timestamps)):
                if values[i] > minacu_cpu:
                    logging.info(f"{timestamps[i].strftime('%Y-%m-%dT%H:%M:%SZ')} - {values[i]}%")
                    p_cnt = p_cnt + 1
                    if p_cnt == 10:
                        break

        asv2_cost_2 = round(min_acu*price_per_acu_hour*730+sum_exceed_mincpu_cost,3)
        #print ("min_acu",min_acu,"min_acu*price_per_acu_hour*730",min_acu*price_per_acu_hour*730,"sum_exceed_mincpu_cost",sum_exceed_mincpu_cost)

        save_cost_1_percent = round((ri_1yr_no - asv2_cost_1)/ri_1yr_no,3)
        save_cost_1_format = "{:.1%}".format(save_cost_1_percent)
        save_cost_2_percent = round((ri_1yr_no - asv2_cost_2)/ri_1yr_no,3)
        save_cost_2_format = "{:.1%}".format(save_cost_2_percent)
        avg_cpu_list.append(avg_cpu_util)

        # 当 save_cost_2_percent 列，当 >0 时，将 1 年 ri 无预付的列以及 asv2 sum 的值分别求和
        #if save_cost_2_percent > 0:
        #    mrr_sum_cost_1ynp=mrr_sum_cost_1ynp+ri_1yr_no
        #    mrr_sum_cost_asv2=mrr_sum_cost_asv2+asv2_cost_2
        #print ("成本对比",mrr_sum_cost_1ynp,mrr_sum_cost_asv2)

        output_result.append(f"{account_id},{rds_region},{instance_id},{engine},{engine_version},{instance_class},{instance_vcpu},{avg_cpu_util},{min_cpu_util},{max_cpu_util},{first_time},{last_time},{od_monthly_cost},{ri_1yr_no},{min_acu},{price_per_acu_hour},{asv2_cost_1},{asv2_cost_2},{save_cost_1_format},{save_cost_2_format}")
        output_result_chart.append(f"{rds_region},{instance_id},{engine},{engine_version},{instance_class},{instance_vcpu},{avg_cpu_util},{min_cpu_util},{max_cpu_util},{first_time},{last_time},{od_price_per_unit},{od_monthly_cost},{noup_1yr_price_per_unit},{ri_1yr_no},{allup_3yr_price_per_unit},{ri_3yr_all},{min_acu},{price_per_acu_hour},{asv2_cost_1},{asv2_cost_2},{save_cost_1_percent},{save_cost_2_percent}")
        #print("output_result",output_result)
        logging.info("-----------------------")
    
    output_column = "account_id,region,instance id,engine,engine_version,instace type,vcpu,CPU Avg Util%,CPU Min Util%,CPU Max Util%,StartTime,EndTime,Ondemand/monthly,1 YR NP/monthly,Min ACU,ASv2 Price/h,ASV2 Cost 1/monthly,ASV2 Cost 2/monthly,Save Percent 1, Save Percent 2"
    logging.info("The evaluation results are as follows. It is recommended to copy and paste them into Excel for reading.")
    logging.info(f"{output_column}")
    
    # 计算 asv2 和 1ynp MRR 成本差异
    #mrr_sum_save= round((mrr_sum_cost_1ynp-mrr_sum_cost_asv2),2)
    #split_columns = output_column.split(',')
    #mrr_column_count = len(split_columns)+2
    #mrr_output_result.append(f"{mrr_sum_cost_1ynp},{mrr_sum_cost_asv2},{mrr_sum_save}")
    #mrr_output_column = "Sum(Ondemand Save Percent 2>0),Sum of(ASV2 Cost 2),ASV2 Cost Savings (Save Percent 2> 0)"

    #top_cpu_output_result = sorted(output_result_chart, key=lambda x: int(x.split(',')[6]), reverse=True)[:8]
    #top_cost_save_output_result = sorted(output_result_chart, key=lambda x: float(x.split(',')[22]), reverse=True)[:8]
    #create_cpu_usage_distribution_chart(count_cpu_usage_distribution(avg_cpu_list), myworksheet, myworkbook)
    #create_top_cpu_chart(top_cpu_output_result,myworksheet,myworkbook)
    #create_top_cost_saving_chart(top_cost_save_output_result,myworksheet,myworkbook)
    #for line in top_cost_save_output_result:
    #    logging.info(line)

    output_df = pd.DataFrame([row.split(',') for row in output_result], columns=output_column.split(','))
    #mrr_output_df = pd.DataFrame([row.split(',') for row in mrr_output_result], columns=mrr_output_column.split(','))

    myworksheet_detail = myworkbook.create_sheet("Detail")
    
    column_names = output_column.split(",")
    df = pd.DataFrame([line.split(",") for line in output_result], columns=column_names)
    myworksheet_detail.append(column_names)
    for row in df.itertuples(index=False):
        myworksheet_detail.append(row)
    
    #mrr_column_names = mrr_output_column.split(",")
    #mrr_df = pd.DataFrame([line.split(",") for line in mrr_output_result], columns=mrr_column_names)
   
    # 写入 mrr_df 的列名，从第 23 列开始，底纹为淡绿色
    #start_col = mrr_column_count
    #header_font = Font(bold=True)
    #light_green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    #for col_idx, col_name in enumerate(mrr_column_names, start=start_col):
    #    cell = myworksheet_detail.cell(row=1, column=col_idx, value=col_name)
    #    cell.font = header_font
    #    cell.fill = light_green_fill

    # 写入 mrr_df 的数据，从第 23 列开始，底纹为淡绿色
    #for row_num, row in enumerate(mrr_df.values, start=2):
    #    for col_idx, value in enumerate(row, start=start_col):
    #        cell = myworksheet_detail.cell(row=row_num, column=col_idx, value=value)
    #        cell.fill = light_green_fill
    # 保存 Excel 文件
    myworkbook.save("asv2_evaluation_report.xlsx")


if __name__ == "__main__":
    try:
        main()
        print("\nProcessing complete!")
    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")
        logging.error(traceback.format_exc())
        logging.info("Please contact the AWS team for processing.")
        raise
