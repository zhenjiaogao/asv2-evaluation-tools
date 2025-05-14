import boto3
import traceback
import math
import json
import os
from datetime import datetime, timedelta
import time
import pandas as pd
import openpyxl
from openpyxl.chart import PieChart, Reference
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Font
import logging
import sys
from operator import itemgetter
import matplotlib.pyplot as plt
import numpy as np
import concurrent.futures
from botocore.exceptions import ClientError

priceList = []
counter = 0

IMG_WIDTH = 600
IMG_HEIGHT = 400
# 设置环境变量
region_list = ['us-east-1','ap-northeast-1','us-east-2','us-west-1','us-west-2','ap-east-1','ap-south-1','ap-southeast-1','ap-northeast-2','ap-southeast-2','ca-central-1','eu-central-1','eu-west-1','eu-west-2','eu-west-3','eu-north-1','me-south-1','sa-east-1']

#region_list = ['us-east-1','ap-northeast-1']

print("Please select a region by entering the corresponding number:")
for i, rg in enumerate(region_list, start=1):
    print(f"{i}. {rg}")
user_input = input(f"Enter your choice (1-{len(region_list)}): ")
if user_input.isdigit() and 1 <= int(user_input) <= len(region_list):
    rds_region = region_list[int(user_input) - 1]
    print(f"You selected: {rds_region}")
else:
    print("Invalid input. Please try again.")

#rds_region = 'us-east-1'
os.environ['AWS_DEFAULT_REGION'] = rds_region

# 创建一个新的 Excel 工作簿
myworkbook = openpyxl.Workbook()

# 获取默认工作表
myworksheet = myworkbook.active
myworksheet.title = "Summary"

# 获取当前日期时间作为日志文件名
log_filename = f"rds_explorer_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

# 配置logging
logging.basicConfig(
    filename=log_filename,
    level=logging.INFO,
    format='%(asctime)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

def get_cpu_utilization(instance_id, region):
    """
    获取给定 RDS 实例在最近一周内的 CPU 使用率统计信息。

    参数:
    instance_id (str): RDS 实例的 ID
    region (str, optional): AWS 区域, 默认为 'us-east-1'
    stat (str, optional): 要获取的 CPU 使用率统计信息, 可选值为 'Average', 'Minimum', 'Maximum', 默认为 'Average'

    返回:
    float: CPU 使用率百分比
    """
    cloudwatch_client = boto3.client('cloudwatch', region_name=rds_region)

    response = cloudwatch_client.get_metric_data(
        MetricDataQueries=[
            {
                'Id': 'cpu_avg',
                'MetricStat': {
                    'Metric': {
                        'Namespace': 'AWS/RDS',
                        'MetricName': 'CPUUtilization',
                        'Dimensions': [
                            {
                                'Name': 'DBInstanceIdentifier',
                                'Value': instance_id
                            }
                        ]
                    },
                    'Period': 86400 * 31,  # 注意：这么大的周期可能会导致数据点不足
                    'Stat': 'Average'
                },
                'ReturnData': True
            },
            {
                'Id': 'cpu_min',
                'MetricStat': {
                    'Metric': {
                        'Namespace': 'AWS/RDS',
                        'MetricName': 'CPUUtilization',
                        'Dimensions': [
                            {
                                'Name': 'DBInstanceIdentifier',
                                'Value': instance_id
                            }
                        ]
                    },
                    'Period': 86400 * 31,
                    'Stat': 'Minimum'
                },
                'ReturnData': True
            },
            {
                'Id': 'cpu_max',
                'MetricStat': {
                    'Metric': {
                        'Namespace': 'AWS/RDS',
                        'MetricName': 'CPUUtilization',
                        'Dimensions': [
                            {
                                'Name': 'DBInstanceIdentifier',
                                'Value': instance_id
                            }
                        ]
                    },
                    'Period': 86400 * 31,
                    'Stat': 'Maximum'
                },
                'ReturnData': True
            },
            {
                'Id': 'cpu_percent',
                'Expression': 'IF(m1>0,m1)',  # 修正：移除 f-string（不需要）
                'Label': 'CPU Watch',
                'ReturnData': True  # 添加缺失的 ReturnData
            },
            {
                'Id': 'm1',
                'MetricStat': {
                    'Metric': {
                        'Namespace': 'AWS/RDS',
                        'MetricName': 'CPUUtilization',
                        'Dimensions': [
                            {
                                'Name': 'DBInstanceIdentifier',
                                'Value': instance_id
                            }
                        ]
                    },
                    'Period': 60,  # 每分钟一个数据点（更合理）
                    'Stat': 'Maximum',
                    'Unit': 'Percent'
                },
                'ReturnData': False
            }
        ],
        # 修正参数名称为 CamelCase
        StartTime=(datetime.utcnow() - timedelta(days=30)).isoformat() + 'Z',
        EndTime=datetime.utcnow().isoformat() + 'Z'
    )

    if response['MetricDataResults']:
        metrics = {}
        for mdr in response['MetricDataResults']:
            metrics[mdr['Id']] = {'timestamps': mdr['Timestamps'], 'values': mdr['Values']}
        return metrics
    else:
        return None

def pricing_get_product(engine, serverless=False, productFamily=None, instance_class=None):
    global priceList
    logging.info(f"查找价格: instance_class={instance_class}, serverless={serverless}, engine={engine}, productFamily={productFamily}")
    for price in priceList:
        if serverless:
            if (price['product']['productFamily'] == productFamily and price['product']['attributes']['databaseEngine'].lower() == engine.lower() and 'IOOptimizedUsage' not in price['product']['attributes'].get('usagetype', '')):
                return price
        else:
            if (price['product']['productFamily'] == 'Database Instance'and price['product']['attributes']['databaseEngine'].lower() == engine.lower() and price['product']['attributes']['instanceType'] == instance_class):
                return price
    logging.error(f'param: engine={engine}, serverless={serverless}, productFamily={productFamily}, instanceType={instance_class}')
    logging.info(f'priceList Dump: {json.dumps(priceList, indent=2)}')
    raise Exception('price not found')

def get_aurora_serverless_acu_price(engine,rds_region):
    """
    Retrieves the price per ACU-Hour for Aurora Serverless.
    Returns:
        float: The price per ACU-Hour for Aurora Serverless.
    """
    ## Q1:把 RDS 的 PG 和 MySQL 都替换为 Aurora PG 和 MySQL 是为了完成 RDS --> Aurora 的成本核算

    # Create a Pricing client 初始化 API

    if engine == "PostgreSQL":
        engine=engine.replace("PostgreSQL","Aurora PostgreSQL")
    if engine == "mysql":
        engine=engine.replace("mysql","Aurora MySQL")
    
    logging.info(f"acu price engine search: {engine}")
    response = pricing_get_product(engine=engine, serverless=True, productFamily='ServerlessV2')
    for offer_term_code, offer_term_data in response["terms"]["OnDemand"].items():
        for price_dimension_key, price_dimension_data in offer_term_data["priceDimensions"].items():
            price_per_unit = float(price_dimension_data["pricePerUnit"]["USD"])
            return price_per_unit        

def get_all_pages(pricing_client, service_code, filters, max_retries=5):
    """
    分页获取 AWS Pricing API 的所有结果
    :param pricing_client: boto3 Pricing 客户端
    :param service_code: 服务代码（如 'AmazonRDS'）
    :param filters: 筛选条件列表
    :param max_retries: 最大重试次数（处理限流）
    :return: 包含所有页结果的列表
    """
    all_results = []
    next_token = None
    retries = 0
    
    while True:
        try:
            # 构建请求参数
            params = {
                'ServiceCode': service_code,
                'Filters': filters
            }
            if next_token:
                params['NextToken'] = next_token
            
            # 执行请求
            response = pricing_client.get_products(**params)
            all_results.extend(response['PriceList'])
            
            # 检查是否有下一页
            if 'NextToken' in response:
                next_token = response['NextToken']
            else:
                break
                
            # 避免过快请求
            time.sleep(0.5)
            
        except ClientError as e:
            if e.response['Error']['Code'] == 'Throttling':
                retries += 1
                if retries > max_retries:
                    logging.error(f"达到最大重试次数: {e}")
                    break
                wait_time = 2 ** retries  # 指数退避：2s, 4s, 8s...
                logging.warning(f"请求被限流，{wait_time}秒后重试 ({retries}/{max_retries})")
                time.sleep(wait_time)
            else:
                logging.error(f"API 请求错误: {e}")
                break
    
    logging.info(f"获取了 {len(all_results)} 条价格数据")
    return all_results

def pricing_get_products_optimized(rds_region):
    """
    获取指定区域的 RDS 价格信息（包括 Serverless 和标准实例）
    :param rds_region: 区域名称（如 'ap-northeast-1'）
    :return: 无（结果存储在全局变量 priceList 中）
    """
    global priceList
    priceList = []  # 重置全局价格列表
    
    # 创建 Pricing 客户端（必须使用 us-east-1）
    pricing_client = boto3.client('pricing', region_name='us-east-1')
    location = aws_region_to_location(rds_region)
    
    # 获取 Serverless V2 价格
    logging.info(f"开始获取 Aurora Serverless V2 ({rds_region}) 的价格")
    serverless_engines = ['Aurora MySQL', 'Aurora PostgreSQL']
    
    for engine in serverless_engines:
        filters = [
            {'Type': 'TERM_MATCH', 'Field': 'productFamily', 'Value': 'ServerlessV2'},
            {'Type': 'TERM_MATCH', 'Field': 'location', 'Value': location},
            {'Type': 'TERM_MATCH', 'Field': 'databaseEngine', 'Value': engine}
        ]
        
        products = get_all_pages(pricing_client, 'AmazonRDS', filters)
        for product in products:
            priceList.append(json.loads(product))
    
    # 获取标准实例价格
    logging.info(f"开始获取标准 RDS 实例 ({rds_region}) 的价格")
    standard_engines = ['Aurora MySQL', 'Aurora PostgreSQL', 'MySql', 'PostgreSQL']
    
    for engine in standard_engines:
        filters = [
            {'Type': 'TERM_MATCH', 'Field': 'databaseEngine', 'Value': engine},
            {'Type': 'TERM_MATCH', 'Field': 'location', 'Value': location},
            {'Type': 'TERM_MATCH', 'Field': 'deploymentOption', 'Value': 'Single-AZ'},
            {'Type': 'TERM_MATCH', 'Field': 'storage', 'Value': 'EBS Only'}
        ]
        
        products = get_all_pages(pricing_client, 'AmazonRDS', filters)
        for product in products:
            priceList.append(json.loads(product))
    
    logging.info(f"完成所有价格获取，共收集 {len(priceList)} 条价格数据")

def count_cpu_usage_distribution(cpu_usage_data):
    """
    统计 CPU 使用率分布情况。

    参数:
    cpu_usage_data (list): 一个包含 CPU 使用率数据的列表。

    返回:
    dict: 一个包含每个 CPU 使用率范围及其对应数量的字典。
    """
    # 定义 CPU 使用率范围
    usage_ranges = [
        ('0% - 10%', 0, 10),
        ('10% - 20%', 10, 20),
        ('20% - 30%', 20, 30),
        ('30% - 50%', 30, 50),
        ('50% and above', 50, 100)
    ]

    # 初始化计数器
    usage_counts = [0] * len(usage_ranges)

    # 统计每个 CPU 使用率范围的数量
    for usage in cpu_usage_data:
        for i, (_, min_range, max_range) in enumerate(usage_ranges):
            if min_range <= usage < max_range:
                usage_counts[i] += 1
                break

    # 返回每个 CPU 使用率范围的数量
    result = [['CPU Usage Range', 'Percentage']]
    result.extend([[range_name, count] for range_name, count in zip(
        [range_name for range_name, _, _ in usage_ranges], usage_counts
    )])
    return result

def create_top_cpu_chart(data, worksheet, workbook):
    # 将数据写入工作表
    id = 2
    worksheet.cell(row=1, column=12, value='instance')
    worksheet.cell(row=1, column=13, value='avg cpu')
    worksheet.cell(row=1, column=14, value='max cpu')
    instance_cpu_data = []
    
    for row in data:
        cells = row.split(',')
        #selected_columns = [0,1, 2, 4,6]
        selected_columns = [1,6,8]
        selected_cell = [cells[i] for i in selected_columns]
        for col, value in enumerate(selected_cell, start=1):
            if col == 2 or col == 3:
                value = int(value)
            worksheet.cell(row=id, column=12 + col - 1, value=value)
        instance_cpu_data = instance_cpu_data + selected_cell
        id = id +1

    instances = []
    avg_cpu_usage = []
    max_cpu_usage = []

    for i in range(0, len(instance_cpu_data), 3):
        instances.append(instance_cpu_data[i])
        avg_cpu_usage.append(int(instance_cpu_data[i+1]))
        max_cpu_usage.append(int(instance_cpu_data[i+2]))

    #创建柱状图的JPG图片
    img = Image(create_jpg_bar_chart(instances,avg_cpu_usage,max_cpu_usage))
    img.width = IMG_WIDTH
    img.height = IMG_HEIGHT

    # 将柱状图添加到工作表
    worksheet.add_image(img, "L16")

    # 保存 Excel 文件
    workbook.save("asv2_evaluation_report.xlsx")

def create_jpg_bar_chart(data1,data2,data3):
    #data = [['instance1', 20, 20], ['instance2', 30, 20], ['instance3', 25, 20], ...]
    # 将数据分离为实例名 AVG CPU  MAX CPU
    instances = data1
    avg_cpu_usage = data2
    max_cpu_usage = data3

    # 创建柱状图
    fig, ax = plt.subplots(figsize=(10, 4))
    bar_width = 0.4
    x = np.arange(len(instances))
    ax.bar(x - bar_width/2, avg_cpu_usage, bar_width, label='Avg CPU')
    ax.bar(x + bar_width/2, max_cpu_usage, bar_width, label='Max CPU')

    # 在柱子上显示数值
    for i, v in enumerate(avg_cpu_usage):
        ax.text(x[i] - bar_width/2, v, str(v), ha='center', va='bottom', fontsize=8)
    for i, v in enumerate(max_cpu_usage):
        ax.text(x[i] + bar_width/2, v, str(v), ha='center', va='bottom', fontsize=8)

    # 设置图表标题和坐标轴标签
    ax.set_title("Instance CPU Usage Top 8", fontsize=16)
    ax.set_xlabel("Instance", fontsize=14)
    ax.set_ylabel("CPU Usage (%)", fontsize=14)

    # 旋转 x 轴标签,以便更好地显示实例名称
    ax.set_xticks(x)
    ax.set_xticklabels(instances, rotation=20, fontsize=10)

    # 添加图例
    ax.legend(loc='upper right')

    # 调整图表边距,以确保 x 轴标签完全显示
    plt.subplots_adjust(bottom=0.3)

    # 保存图片为 JPG 格式
    file_name = "instance_avg_cpu_jpg_name.jpg"
    plt.savefig(file_name, dpi=300, bbox_inches='tight')
    return file_name

def create_top_cost_saving_chart(data, worksheet, workbook):
    # 将数据写入工作表
    id=2
    worksheet.cell(row=1, column=23, value='instance')
    worksheet.cell(row=1, column=24, value='Privision 1YR RI NP')
    worksheet.cell(row=1, column=25, value='ASv2 Cost 2/monthly')
    worksheet.cell(row=1, column=26, value='Save Percent 2')

    instance_cost_data = []
    for row in data:
        cells = row.split(',')
        selected_columns = [1,14,20,22]
        selected_cell = [cells[i] for i in selected_columns]
        for col, value in enumerate(selected_cell, start=1):
            if col == 2 or col == 3 or col ==4:
                value = float(value)
            worksheet.cell(row=id, column=23 + col - 1, value=value)
        instance_cost_data = instance_cost_data + selected_cell
        id = id +1

    instances = []
    ri_1year = []
    asv2_cost_2 = []
    asv2_2_saving_percent = []

    for i in range(0, len(instance_cost_data), 4):
        instances.append(instance_cost_data[i])
        asv2_cost_2.append(float(instance_cost_data[i+1]))
        ri_1year.append(float(instance_cost_data[i+2]))
        asv2_2_saving_percent.append(float(instance_cost_data[i+3]))
        #print (instances,asv2_cost_2,ri_1year,asv2_2_saving_percent)

    #创建柱状图的JPG图片
    img = Image(create_cost_bar_jpg(instances,asv2_cost_2,ri_1year))
    img.width = IMG_WIDTH
    img.height = IMG_HEIGHT


    # 将柱状图添加到工作表
    worksheet.add_image(img, "W16")

    # 保存 Excel 文件
    workbook.save("asv2_evaluation_report.xlsx")

def create_cost_bar_jpg(data1,data2,data3):
    #data = [['instance1', 20, 20], ['instance2', 30, 20], ['instance3', 25, 20], ...]
    # 将数据分离为实例名 EstCost  RI1Year
    instances = data1
    asv2_cost_2 = data2
    ri_1year = data3

    # 创建柱状图
    fig, ax = plt.subplots(figsize=(10, 4))
    bar_width = 0.4
    x = np.arange(len(instances))
    ax.bar(x - bar_width/2, asv2_cost_2, bar_width, label='Est Cost')
    ax.bar(x + bar_width/2, ri_1year, bar_width, label='1 Year RI')

    # 在柱子上显示数值
    for i, v in enumerate(asv2_cost_2):
        ax.text(x[i] - bar_width/2, v, str(v), ha='center', va='bottom', fontsize=8)
    for i, v in enumerate(ri_1year):
        ax.text(x[i] + bar_width/2, v, str(v), ha='center', va='bottom', fontsize=8)

    # 设置图表标题和坐标轴标签
    ax.set_title("Instance Cost Saving Top 8", fontsize=16)
    ax.set_xlabel("Instance", fontsize=14)
    ax.set_ylabel("Cost", fontsize=14)

    # 旋转 x 轴标签,以便更好地显示实例名称
    ax.set_xticks(x)
    ax.set_xticklabels(instances, rotation=20, fontsize=10)

    # 添加图例
    ax.legend(loc='upper right')

    # 调整图表边距,以确保 x 轴标签完全显示
    plt.subplots_adjust(bottom=0.3)

    # 保存图片为 JPG 格式
    file_name = "instance_cost_saving_jpg_name.jpg"
    plt.savefig(file_name, dpi=300, bbox_inches='tight')
    return file_name

def create_cpu_usage_distribution_chart(data, worksheet, workbook):
    """
    创建 CPU 使用率分布饼图并将其添加到 Excel 工作表中。

    参数:
    data (list): 包含 CPU 使用率分布数据的列表,第一行为列名。
    worksheet (Worksheet): 要添加图表的 Excel 工作表。
    workbook (Workbook): 包含工作表的 Excel 工作簿。
    """
    # 将数据写入工作表
    for row in data:
        worksheet.append(row)

    # 创建一个饼图
    #chart = PieChart()
    #chart.title = "Instance Count by CPU Avg Utilization"
    #chart.height = 12
    #chart.width = 18

    # 设置图表数据范围
    #labels = Reference(worksheet, min_row=2, max_row=len(data), min_col=1)
    #data_ref = Reference(worksheet, min_row=1, max_row=len(data), min_col=2)
    #chart.add_data(data_ref, titles_from_data=True)
    #chart.set_categories(labels)

    # 添加数字百分比标签
    #s = chart.series[0]
    #s.dLbls = DataLabelList()
    #s.dLbls.showVal = True

    # 创建饼图
    labels = [row[0] for row in data[1:] if float(row[1]) > 0]
    values = [float(row[1]) for row in data[1:] if float(row[1]) > 0]

    fig, ax = plt.subplots(figsize=(8, 8))
    patches, texts, autotexts = ax.pie(values, labels=labels, autopct='%1.1f%%')
    # 设置标签字体大小和颜色
    for t in texts:
        t.set_size('smaller')
        t.set_color('black')

    # 设置百分比字体大小和颜色
    for t in autotexts:
        t.set_size('smaller')
        t.set_color('white')

    ax.axis('equal')  # 确保饼图是圆形的
    ax.set_title("Instance Count by CPU Avg Utilization")

    # 将图像保存为 JPG 格式
    plt.savefig("cpu_usage_pie.jpg", dpi=300)

    # 将 JPG 图像插入到 Excel 工作表的 A16 单元格
    img = Image("cpu_usage_pie.jpg")
    img.width = IMG_WIDTH
    img.height = IMG_HEIGHT
    worksheet.add_image(img, "A16")

    # 保存 Excel 文件
    workbook.save("asv2_evaluation_report.xlsx")

def update_progress(current_step, total_steps):
    """
    更新进度条
    """
    progress = current_step / total_steps * 100
    bar_length = 30  # 进度条长度
    block = int(round(bar_length * progress / 100))
    text = "\rProcessing: [{0}] {1}%".format("#" * block + "-" * (bar_length - block), round(progress, 2))
    sys.stdout.write(text)
    sys.stdout.flush()

def aws_region_to_location(region):
    region_to_location_map = {
        "us-east-1": "US East (N. Virginia)",
        "us-east-2": "US East (Ohio)",
        "us-west-1": "US West (N. California)",
        "us-west-2": "US West (Oregon)",
        "ap-east-1": "Asia Pacific (Hong Kong)",
        "ap-south-1": "Asia Pacific (Mumbai)",
        "ap-northeast-1": "Asia Pacific (Tokyo)",
        "ap-northeast-2": "Asia Pacific (Seoul)",
        "ap-southeast-1": "Asia Pacific (Singapore)",
        "ap-southeast-2": "Asia Pacific (Sydney)",
        "ca-central-1": "Canada (Central)",
        "eu-central-1": "EU (Frankfurt)",
        "eu-west-1": "EU (Ireland)",
        "eu-west-2": "EU (London)",
        "eu-west-3": "EU (Paris)",
        "eu-north-1": "EU (Stockholm)",
        "me-south-1": "Middle East (Bahrain)",
        "sa-east-1": "South America (São Paulo)"
    }

    if region in region_to_location_map:
        return region_to_location_map[region]
    else:
        return "Unknown location"

def process_instance(instance, instance_count):
    global counter
    avg_cpu_list=[]
    counter += 1
    update_progress(counter, instance_count)
    logging.info("-----------------------")
    instance_id = instance['DBInstanceIdentifier']
    instance_class = instance['DBInstanceClass']
    engine = instance['Engine'].replace("-", " ", 1)
    engine_version = instance['EngineVersion']
    status = instance['DBInstanceStatus']
    availability_zone = instance['AvailabilityZone']
    db_cluster_identifier = instance.get('DBClusterIdentifier', '')
    account_id = instance['DBInstanceArn'].split(':')[4]

    if engine == "postgres":
        engine = engine.replace("postgres", "PostgreSQL")

    product_json = pricing_get_product(engine=engine, serverless=False, instance_class=instance_class)
    vcpu = product_json['product']['attributes']['vcpu']

    # 获取OD单价
    for offer_term_code, offer_term_data in product_json['terms']["OnDemand"].items():
        for price_dimension_code, price_dimension_data in offer_term_data["priceDimensions"].items():
            od_price_per_unit = round(float(price_dimension_data["pricePerUnit"]["USD"]), 3)
            logging.info(f"OD Price per unit: {od_price_per_unit}")

    # 获取1 year no upfront的单价
    if 'Reserved' in product_json['terms']:
        for offer_term_code, offer_term_data in product_json['terms']["Reserved"].items():
            if offer_term_data["termAttributes"]["LeaseContractLength"] == "1yr" and offer_term_data["termAttributes"][
                "PurchaseOption"] == "No Upfront":
                for price_dimension_code, price_dimension_data in offer_term_data["priceDimensions"].items():
                    noup_1yr_price_per_unit = round(float(price_dimension_data["pricePerUnit"]["USD"]), 3)
                    logging.info(f"No upfront 1 year Price per unit: {noup_1yr_price_per_unit}")
    else:
        logging.info("'Reserved' node not found in product_json['terms']")
        noup_1yr_price_per_unit = 0

    # 获取3 year all upfront的单价
    if 'Reserved' in product_json['terms']:
        for offer_term_code, offer_term_data in product_json['terms']["Reserved"].items():
            if offer_term_data["termAttributes"]["LeaseContractLength"] == "3yr" and offer_term_data["termAttributes"][
                "PurchaseOption"] == "All Upfront":
                for price_dimension_code, price_dimension_data in offer_term_data["priceDimensions"].items():
                    if price_dimension_data["unit"] == "Quantity":
                        allup_3yr_price_per_unit = round(float(price_dimension_data["pricePerUnit"]["USD"]) / 36 / 720,
                                                         3)
                        logging.info(f"All upfront 3 year Price per unit: {allup_3yr_price_per_unit}")
    else:
        logging.info("'Reserved' node not found in product_json['terms']")
        allup_3yr_price_per_unit = 0

    # 获取acu的单价
    price_per_acu_hour = get_aurora_serverless_acu_price(engine,rds_region)
    logging.info(f"Price per ACU-Hour: {price_per_acu_hour}")

    # 获取最近1月的CPU 使用率情况
    cpu_utils = get_cpu_utilization(instance_id, rds_region)  # get avg/min/max and percent in one shot
    avg_cpu_util = math.ceil(cpu_utils.get('cpu_avg')['values'][0])
    if avg_cpu_util is not None:
        logging.info(f"Average CPU Utilization (1 week): {avg_cpu_util:.2f}%")
    else:
        logging.info("No CPU utilization data available")

    min_cpu_util = math.ceil(cpu_utils.get('cpu_min')['values'][0])
    if min_cpu_util is not None:
        logging.info(f"Minimum CPU Utilization (1 week): {min_cpu_util:.2f}%")
    else:
        logging.info("No CPU utilization data available")

    max_cpu_util =  math.ceil(cpu_utils.get('cpu_max')['values'][0])
    if max_cpu_util is not None:
        logging.info(f"Maximum CPU Utilization (1 week): {max_cpu_util:.2f}%")
    else:
        logging.info("No CPU utilization data available")

    od_monthly_cost = round(730*od_price_per_unit,3)
    ri_1yr_no = round(730*noup_1yr_price_per_unit,3)
    ri_3yr_all = round(730*allup_3yr_price_per_unit,3)
        
    ## Q：下面的公式要除以 100 是为了将利用率转换为正常可计算的数值
    #print ("before min acu",(avg_cpu_util+min_cpu_util)/100/2*int(vcpu)*4)
    before_min_acu = (avg_cpu_util+min_cpu_util)/100/2*int(vcpu)*4
    if before_min_acu <= 0.5:
        min_acu = 0.5
    else:
        min_acu = math.ceil((avg_cpu_util+min_cpu_util)/100/2*int(vcpu)*4) 
    #print ("min_acu",min_acu)
    ## Q: 同上
    before_avg_acu=avg_cpu_util/100*int(vcpu)*4
    if before_avg_acu <=0.5:
        avg_acu = 0.5
    else:
        avg_acu = math.ceil(avg_cpu_util/100*int(vcpu)*4)
    #print('avg acu',avg_acu,"before_avg_acu",before_avg_acu,"avg_cpu_util",avg_cpu_util)
    asv2_cost_1 = avg_acu*730*price_per_acu_hour
    sum_exceed_mincpu_cost = 0

    # 获取CPU 使用率超过MinACU的数据(=CPU%>(AVG+MIN)/2)
    minacu_cpu=(avg_cpu_util+min_cpu_util)/2
    high_cpu_data = cpu_utils['cpu_percent']
    if high_cpu_data:
        timestamps = high_cpu_data['timestamps']
        values = high_cpu_data['values']
        timestamps.reverse()
        first_time = timestamps[0]
        last_time = timestamps[-1]
        values.reverse()
        all_values = values
        exceed_minacu_value_cnt = 0
        logging.info(f"all values cnt : {len(all_values)}")

        # 计算弹性费用，超过MinACU的情况
        for i in range(len(timestamps)):
            if values[i] > minacu_cpu:
                exceed_mincpu_cost = (math.ceil(values[i] / 100 * int(vcpu) * 4) - min_acu) * price_per_acu_hour / 60
                sum_exceed_mincpu_cost = sum_exceed_mincpu_cost + exceed_mincpu_cost
                exceed_minacu_value_cnt = exceed_minacu_value_cnt + 1

        logging.info(f"CPU Utilization > {minacu_cpu}%: total {exceed_minacu_value_cnt}, only logging.info 10 samples")
        p_cnt = 0
        for i in range(len(timestamps)):
            if values[i] > minacu_cpu:
                logging.info(f"{timestamps[i].strftime('%Y-%m-%dT%H:%M:%SZ')} - {values[i]}%")
                p_cnt = p_cnt + 1
                if p_cnt == 10:
                    break
        
        asv2_cost_2 = round(min_acu*price_per_acu_hour*730+sum_exceed_mincpu_cost,3)
        #print ("min_acu",min_acu,"min_acu*price_per_acu_hour*730",min_acu*price_per_acu_hour*730,"sum_exceed_mincpu_cost",sum_exceed_mincpu_cost)

        save_cost_1_percent = round((ri_1yr_no - asv2_cost_1)/ri_1yr_no,3)
        save_cost_1_format = "{:.1%}".format(save_cost_1_percent)
        save_cost_2_percent = round((ri_1yr_no - asv2_cost_2)/ri_1yr_no,3)
        save_cost_2_format = "{:.1%}".format(save_cost_2_percent)
        avg_cpu_list.append(avg_cpu_util)
    # return avg_cpu_util, output_result, output_result_chart
    return (avg_cpu_util,
            f"{account_id},{rds_region},{instance_id},{engine},{engine_version},{instance_class},{vcpu},{avg_cpu_util},{min_cpu_util},{max_cpu_util},{first_time},{last_time},{od_monthly_cost},{ri_1yr_no},{ri_3yr_all},{min_acu},{price_per_acu_hour},{asv2_cost_1},{asv2_cost_2},{save_cost_1_format},{save_cost_2_format}",
            f"{rds_region},{instance_id},{engine},{engine_version},{instance_class},{vcpu},{avg_cpu_util},{min_cpu_util},{max_cpu_util},{first_time},{last_time},{od_price_per_unit},{od_monthly_cost},{noup_1yr_price_per_unit},{ri_1yr_no},{allup_3yr_price_per_unit},{ri_3yr_all},{min_acu},{price_per_acu_hour},{asv2_cost_1},{asv2_cost_2},{save_cost_1_percent},{save_cost_2_percent}")

def main():
    output_result=[]
    avg_cpu_list=[]
    output_result_chart=[]
    rds = boto3.client('rds')
    marker = None
    all_filtered_instances = []

    while True:
        if marker:
            instance_response = rds.describe_db_instances(Marker=marker)
        else:
            instance_response = rds.describe_db_instances()
        
        instances = instance_response['DBInstances']
        filtered = [
            instance for instance in instances
            if instance['Engine'] in ['aurora-mysql', 'aurora-postgresql', 'mysql', 'postgres']
            and instance['DBInstanceClass'] != 'db.serverless'
        ]
        all_filtered_instances.extend(filtered)
        
        if 'Marker' in instance_response:
            marker = instance_response['Marker']
        else:
            break

    # 一次获取该 region 下的 rds 及 aurora 所有产品的 price
    pricing_get_products_optimized(rds_region)

    # 分批处理实例（每批50个）
    batch_size = 2
    total_instances = len(all_filtered_instances)
    total_batches = (total_instances + batch_size - 1) // batch_size  # 向上取整

    print(f"######## 总批次为 {total_batches}, 共计 {total_instances} 个实例 ########")
    # 处理每一批实例
    for batch_idx in range(total_batches):
        start_idx = batch_idx * batch_size
        end_idx = min(start_idx + batch_size, total_instances)
        batch_instances = all_filtered_instances[start_idx:end_idx]
        print(f"######## 开始处理第 {batch_idx+1}/{total_batches} 批次, 实例 {start_idx+1} ~ {end_idx}/{total_instances} ########")
        
        # 使用50个并发线程处理当前批次的实例
        results = []
        with concurrent.futures.ThreadPoolExecutor(max_workers=1) as executor:
            tasks = [
                executor.submit(process_instance, instance, total_instances)
                for instance in batch_instances
            ]
            
            # 等待所有任务完成，并处理结果
            for future in concurrent.futures.as_completed(tasks):
                try:
                    result = future.result()
                    results.append(result)
                except Exception as e:
                    print(f"处理实例时出错: {e}")
        
        # 可以在这里处理当前批次的结果（如保存到文件）
        print(f"批次 {batch_idx+1} 处理完成，成功处理 {len(results)} 个实例")

    # 处理数据
    for result in results:
        avg_cpu_list.append(result[0])
        output_result.append(result[1])
        output_result_chart.append(result[2])
    output_column = "account_id,region,instance id,engine,engine_version,instace type,vcpu,CPU Avg Util%,CPU Min Util%,CPU Max Util%,StartTime,EndTime,Ondemand/monthly,1 YR NP/monthly,3 YR AP/monthly,Min ACU,ASv2 Price/h,ASV2 Cost 1/monthly,ASV2 Cost 2/monthly,Save Percent 1, Save Percent 2"
    logging.info(
        "The evaluation results are as follows. It is recommended to copy and paste them into Excel for reading.")
    logging.info(f"{output_column}")

    # logger.info(f'debug for output_result_chart: {json.dumps(output_result_chart, indent=2)}')
    top_cpu_output_result = sorted(output_result_chart, key=lambda x: float(x.split(',')[6]), reverse=True)[:8]
    top_cost_save_output_result = sorted(output_result_chart, key=lambda x: float(x.split(',')[22]), reverse=True)[:8]
    create_cpu_usage_distribution_chart(count_cpu_usage_distribution(avg_cpu_list), myworksheet, myworkbook)
    create_top_cpu_chart(top_cpu_output_result, myworksheet, myworkbook)
    create_top_cost_saving_chart(top_cost_save_output_result, myworksheet, myworkbook)
    for line in top_cost_save_output_result:
        logging.info(line)

    output_df = pd.DataFrame([row.split(',') for row in output_result], columns=output_column.split(','))

    myworksheet_detail = myworkbook.create_sheet("Detail")

    column_names = output_column.split(",")
    df = pd.DataFrame([line.split(",") for line in output_result], columns=column_names)
    myworksheet_detail.append(column_names)
    for row in df.itertuples(index=False):
        myworksheet_detail.append(row)

    # 结果保存到 xlsx 文件中
    myworkbook.save("asv2_evaluation_report.xlsx")

if __name__ == "__main__":
    try:
        main()
        print("\nProcessing complete!")
    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")
        logging.error(traceback.format_exc())
        logging.info("Please contact the AWS team for processing.")
        raise
